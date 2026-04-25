#!/usr/bin/env python3
"""
exportar_cargas.py
Lee DIESEL NISSAN Y TALLER *.xlsm y genera un .xlsx plano importable por el sistema WB.
Extrae cargas de patio (cols D-I) y campo (cols T-AA) de la hoja CARGAS GENERAL.

Uso:
  python3 exportar_cargas.py                        # procesa todo el historial
  python3 exportar_cargas.py 2026-04-04 2026-04-10  # filtra por rango de fechas
"""

import re
import sys
import openpyxl
from openpyxl import Workbook
from datetime import date

# ── Configuración ─────────────────────────────────────────────────────────────

INPUT  = "DIESEL NISSAN Y TALLER 04 al 10  de ABRIL DE  2026 v1.xlsm"
OUTPUT = "cargas_limpias.xlsx"

# Columnas (índice 0 = col A)
# Patio / Taller (sección izquierda)
C_PATIO_HORA      = 3   # D
C_PATIO_FOLIO     = 4   # E
C_PATIO_OPERADOR  = 5   # F  (Responsable)
C_PATIO_UNIDAD    = 6   # G
C_PATIO_LITROS    = 7   # H  (cantidad)
C_PATIO_ODOMETRO  = 8   # I  (HRS/KM)

# Campo / Nissan externas (sección derecha)
C_CAMPO_HORA      = 19  # T
C_CAMPO_FOLIO     = 20  # U
C_CAMPO_UNIDAD    = 21  # V
C_CAMPO_LITROS    = 22  # W
C_CAMPO_ODOMETRO  = 24  # Y  (HRS/KM)
C_CAMPO_OBRA      = 25  # Z
C_CAMPO_OPERADOR  = 26  # AA

# Meses en español
MESES = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4,
    "MAYO": 5, "JUNIO": 6, "JULIO": 7, "AGOSTO": 8,
    "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12,
}

DIAS = ["LUNES", "MARTES", "MIERCOLES", "MIÉRCOLES", "JUEVES",
        "VIERNES", "SABADO", "SÁBADO", "DOMINGO"]

# Palabras que identifican filas que NO son cargas reales
SKIP_WORDS = [
    "SUMINISTRADO", "DISPONIBLE", "CARGAS", "DIESEL INICIO",
    "FOLIO", "INICIO JORNADA", "FIN JORNADA",
    "AJUSTE", "MAX ", "VACIO", "CANTIDAD", "LITROS",
    "HORA", "UNIDAD", "RESPONSABLE",
]


# ── Helpers ───────────────────────────────────────────────────────────────────

def parse_fecha_header(texto: str) -> date | None:
    """'SABADO 27 DE DICIEMBRE DE 2025' → date(2025, 12, 27)"""
    t = texto.upper().strip()
    for dia in DIAS:
        t = t.replace(dia, "")
    t = re.sub(r"\bDE\b", "", t).strip()
    partes = t.split()
    if len(partes) >= 3:
        try:
            dia_num = int(partes[0])
            mes     = MESES.get(partes[1], 0)
            anio    = int(partes[2])
            if mes and 1 <= dia_num <= 31 and 2020 <= anio <= 2030:
                return date(anio, mes, dia_num)
        except (ValueError, IndexError):
            pass
    return None


def es_header_fecha(val) -> bool:
    if not isinstance(val, str):
        return False
    return any(d in val.upper() for d in DIAS)


def es_fila_skip(val) -> bool:
    if not isinstance(val, str):
        return False
    v = val.upper()
    return any(k in v for k in SKIP_WORDS)


def parse_str(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    if s == "" or s.startswith("="):
        return None
    return s


def parse_num(val) -> float | None:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if s.startswith("="):
        return None
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return None


def parse_hora(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    # "08:00:00" o "8:00:00"
    m = re.match(r"^(\d{1,2}):(\d{2})", s)
    if m:
        return f"{int(m.group(1)):02d}:{m.group(2)}"
    # Serial de Excel como float
    if isinstance(val, float) and 0 < val < 1:
        total_min = round(val * 1440)
        return f"{total_min // 60:02d}:{total_min % 60:02d}"
    return None


def unidad_valida(val) -> bool:
    s = parse_str(val)
    if not s:
        return False
    if es_fila_skip(s):
        return False
    # Códigos de unidad siempre tienen letras (CA12, EX08, NISSAN…)
    # Números puros son folios que se cuelan en la columna V
    if not re.search(r"[A-Za-z]", s):
        return False
    return True


# ── Parser principal ──────────────────────────────────────────────────────────

def extraer_cargas(rango_inicio: date | None, rango_fin: date | None) -> list[dict]:
    print(f"Abriendo {INPUT} ...")
    wb = openpyxl.load_workbook(INPUT, read_only=True, keep_vba=True, data_only=True)

    if "CARGAS GENERAL" not in wb.sheetnames:
        sys.exit("ERROR: No se encontró la hoja 'CARGAS GENERAL'")

    ws   = wb["CARGAS GENERAL"]
    rows = []
    fecha_actual: date | None = None

    for fila_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not row:
            continue

        # Padding para evitar IndexError en filas cortas
        r = list(row) + [None] * 10

        val_d = r[C_PATIO_HORA]   # col D

        # ¿Es encabezado de fecha?
        if es_header_fecha(str(val_d or "")):
            parsed = parse_fecha_header(str(val_d))
            if parsed:
                fecha_actual = parsed
            continue

        if fecha_actual is None:
            continue

        # Filtro por rango
        if rango_inicio and fecha_actual < rango_inicio:
            continue
        if rango_fin and fecha_actual > rango_fin:
            continue

        fecha_str = fecha_actual.isoformat()

        # ── Patio ────────────────────────────────────────────────
        p_unidad = parse_str(r[C_PATIO_UNIDAD])
        p_litros = parse_num(r[C_PATIO_LITROS])

        if unidad_valida(p_unidad) and p_litros and p_litros > 0:
            rows.append({
                "fecha":       fecha_str,
                "hora":        parse_hora(r[C_PATIO_HORA]),
                "folio":       int(v) if (v := parse_num(r[C_PATIO_FOLIO])) else "",
                "unidad":      p_unidad.upper(),
                "litros":      p_litros,
                "origen":      "patio",
                "operador":    parse_str(r[C_PATIO_OPERADOR]) or "",
                "obra":        "",
                "odometro_hrs": parse_num(r[C_PATIO_ODOMETRO]) or "",
                "tipo_diesel": "normal",
            })

        # ── Campo ────────────────────────────────────────────────
        c_unidad = parse_str(r[C_CAMPO_UNIDAD])
        c_litros = parse_num(r[C_CAMPO_LITROS])

        if unidad_valida(c_unidad) and c_litros and c_litros > 0:
            rows.append({
                "fecha":       fecha_str,
                "hora":        parse_hora(r[C_CAMPO_HORA]),
                "folio":       int(v) if (v := parse_num(r[C_CAMPO_FOLIO])) else "",
                "unidad":      c_unidad.upper(),
                "litros":      c_litros,
                "origen":      "campo",
                "operador":    parse_str(r[C_CAMPO_OPERADOR]) or "",
                "obra":        parse_str(r[C_CAMPO_OBRA]) or "",
                "odometro_hrs": parse_num(r[C_CAMPO_ODOMETRO]) or "",
                "tipo_diesel": "normal",
            })

    return rows


# ── Salida Excel ──────────────────────────────────────────────────────────────

def escribir_xlsx(rows: list[dict], output: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "CARGAS GENERAL"

    headers = [
        "fecha", "hora", "folio", "unidad", "litros",
        "origen", "operador", "obra", "odometro_hrs", "tipo_diesel",
    ]
    ws.append(headers)

    for r in rows:
        ws.append([
            r["fecha"], r["hora"], r["folio"], r["unidad"], r["litros"],
            r["origen"], r["operador"], r["obra"], r["odometro_hrs"], r["tipo_diesel"],
        ])

    wb.save(output)


# ── Resumen ───────────────────────────────────────────────────────────────────

def imprimir_resumen(rows: list[dict]):
    if not rows:
        print("No se extrajeron filas.")
        return

    fechas  = sorted(set(r["fecha"] for r in rows))
    patio   = [r for r in rows if r["origen"] == "patio"]
    campo   = [r for r in rows if r["origen"] == "campo"]
    unidades = sorted(set(r["unidad"] for r in rows))

    print(f"\n{'─'*50}")
    print(f"  Total cargas    : {len(rows)}")
    print(f"  Patio (taller)  : {len(patio)}")
    print(f"  Campo (Nissan)  : {len(campo)}")
    print(f"  Rango de fechas : {fechas[0]}  →  {fechas[-1]}")
    print(f"  Días con datos  : {len(fechas)}")
    print(f"  Unidades únicas : {len(unidades)}")
    print(f"  Litros totales  : {sum(r['litros'] for r in rows):,.0f} L")
    print(f"{'─'*50}")

    # Primeras 10 filas como muestra
    print("\nMuestra (primeras 10 filas):")
    print(f"  {'fecha':<12} {'unidad':<14} {'litros':>7} {'origen':<7} {'folio':>7}  operador")
    for r in rows[:10]:
        print(f"  {r['fecha']:<12} {r['unidad']:<14} {r['litros']:>7.0f} {r['origen']:<7} {str(r['folio']):>7}  {r['operador']}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    rango_inicio: date | None = None
    rango_fin:    date | None = None

    if len(sys.argv) >= 3:
        try:
            rango_inicio = date.fromisoformat(sys.argv[1])
            rango_fin    = date.fromisoformat(sys.argv[2])
            print(f"Filtrando: {rango_inicio} → {rango_fin}")
        except ValueError:
            sys.exit("Uso: python3 exportar_cargas.py [YYYY-MM-DD] [YYYY-MM-DD]")

    rows = extraer_cargas(rango_inicio, rango_fin)
    imprimir_resumen(rows)

    if rows:
        escribir_xlsx(rows, OUTPUT)
        print(f"\nArchivo generado: {OUTPUT}")
        print("Ahora puedes subirlo en Admin → Importar del sistema.")
    else:
        print("No se generó archivo (sin datos).")


if __name__ == "__main__":
    main()
